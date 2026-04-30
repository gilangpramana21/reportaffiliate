"""
DataParser — membaca file spreadsheet (Excel/CSV/ODS) dan menghasilkan ParseResult
berisi deal_rows dan non_deal_rows sebagai list CreatorRow.
"""
from __future__ import annotations

import math
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, TYPE_CHECKING

import pandas as pd
from .template_detector import TemplateDetector, TemplateInfo

if TYPE_CHECKING:
    from .multi_brand_detector import BrandDetectionResult


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------

class UnsupportedFormatError(Exception):
    """Raised when the uploaded file format is not supported."""


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class CreatorRow:
    tanggal: Optional[str] = None
    username: Optional[str] = None
    link_acc: Optional[str] = None
    followers: Optional[int] = None
    contact: Optional[str] = None
    brand: Optional[str] = None
    pic: Optional[str] = None
    avg_gmv_month: Optional[float] = None
    gmv_per_pembeli: Optional[float] = None
    update: Optional[str] = None
    respon_speed: Optional[str] = None
    result: Optional[str] = None
    sampel_gratis: Optional[str] = None
    note: Optional[str] = None
    total_vt: Optional[int] = None          # Total jumlah video yang diupload
    note_deal: Optional[str] = None         # Catatan deal
    gmv_perbulan: Optional[float] = None    # GMV per bulan after join
    video_links: list = field(default_factory=list)
    custom_fields: dict = field(default_factory=dict)


@dataclass
class ParseResult:
    deal_rows: list
    non_deal_rows: list
    detected_columns: list
    errors: list
    template_info: Optional[TemplateInfo] = None  # Informasi template yang terdeteksi
    # Multi-brand support fields
    brand_detection_result: Optional['BrandDetectionResult'] = None
    is_multi_brand: bool = False


# ---------------------------------------------------------------------------
# DataParser
# ---------------------------------------------------------------------------

class DataParser:
    SUPPORTED_FORMATS = ['.xlsx', '.xls', '.csv', '.ods']
    SHEET_DEAL = 'Deal'
    SHEET_NON_DEAL = 'Belum Deal'

    def __init__(self):
        self.template_detector = TemplateDetector()

    # Kolom standar yang dikenali (uppercase, digunakan untuk deteksi header)
    STANDARD_COLUMNS = [
        'TANGGAL', 'TGL DEAL', 'TANGGAL DEAL', 'USERNAME', 'LINK ACC', 'FOLLS', 'CONTACT',
        'BRAND', 'PIC', 'AVG GMV/MONTH', 'GMV PER PEMBELI',
        'UPDATE', 'RESPON SPEED', 'RESPON', 'SPEED', 'RESULT', 'SAMPEL GRATIS',
        'STATUS SEMPEL', 'NOTE', 'LINK VIDEO',
        # Kolom tambahan
        'TOTAL VT', 'NOTE DEAL', 'NOTE DARI RARA', 'NOTE RARA',
        'GMV PERBULAN AFTER JOIN', 'GMV PER WEEK 1 AFER JOIN',
        'GMV PER WEEK 2 AFER JOIN', 'GMV PER WEEK 3 AFER JOIN',
        'GMV PER WEEK 4 AFER JOIN',
        'APPROVED BY RARA',
        # Alias Florist & variasi lain
        'TANGGAL DEAL', 'FOLLOWERS', 'FOLLOWER', 'LINK AKUN', 'LINK ACCOUNT',
        'SAMPEL APA', 'UPDATE VT', 'NOTE DEAL VT', 'JUMLAH VT',
        'TOTAL GMV', 'TOTAL GMV WEEK', 'AVG GMV',
        # Variasi kolom link video dengan followup
        'UPDATE VT FOLLOWUP', 'LINK VIDEO FOLLOWUP', 'VIDEO LINK',
    ]

    # Mapping nama kolom standar → field CreatorRow (termasuk alias)
    _COLUMN_TO_FIELD = {
        'TANGGAL': 'tanggal',
        'TGL DEAL': 'tanggal',
        'TANGGAL DEAL': 'tanggal',
        'USERNAME': 'username',
        'LINK ACC': 'link_acc',
        'LINK AKUN': 'link_acc',
        'LINK ACCOUNT': 'link_acc',
        'FOLLS': 'followers',
        'FOLLOWERS': 'followers',
        'FOLLOWER': 'followers',
        'CONTACT': 'contact',
        'BRAND': 'brand',
        'PIC': 'pic',
        'APPROVED BY RARA': 'pic',
        'AVG GMV/MONTH': 'avg_gmv_month',
        'AVG GMV/\nMONTH': 'avg_gmv_month',
        'AVG GMV/ MONTH': 'avg_gmv_month',
        'AVG GMV': 'avg_gmv_month',
        'GMV PER PEMBELI': 'gmv_per_pembeli',
        'UPDATE': 'update',
        'UPDATE VT': 'update',
        'RESPON SPEED': 'respon_speed',
        'RESPON': 'respon_speed',
        'SPEED': 'respon_speed',
        'RESULT': 'result',
        'SAMPEL GRATIS': 'sampel_gratis',
        'STATUS SEMPEL': 'sampel_gratis',
        'SAMPEL APA': 'sampel_gratis',
        'NOTE': 'note',
        'NOTE DEAL': 'note_deal',
        'NOTE DEAL VT': 'note_deal',
        'NOTE DARI RARA': 'note_deal',
        'NOTE RARA': 'note_deal',
        'LINK VIDEO': 'update',  # Kolom link video tidak perlu mapping khusus, diambil dari cell notes
        'UPDATE VT FOLLOWUP': 'update',
        'LINK VIDEO FOLLOWUP': 'update',
        'VIDEO LINK': 'update',
        'TOTAL VT': 'total_vt',
        'JUMLAH VT': 'total_vt',
        'GMV PERBULAN AFTER JOIN': 'gmv_perbulan',
        'GMV PER WEEK 1 AFER JOIN': 'gmv_perbulan',
        'GMV PER WEEK 2 AFER JOIN': 'gmv_perbulan',
        'GMV PER WEEK 3 AFER JOIN': 'gmv_perbulan',
        'GMV PER WEEK 4 AFER JOIN': 'gmv_perbulan',
        'TOTAL GMV': 'gmv_perbulan',
    }

    # Field yang perlu konversi numerik
    _INT_FIELDS = {'followers', 'total_vt'}
    _FLOAT_FIELDS = {'avg_gmv_month', 'gmv_per_pembeli', 'gmv_perbulan'}

    def detect_multi_brand(self, parse_result: ParseResult) -> ParseResult:
        """
        Perform multi-brand detection on a ParseResult and update it with brand information.
        
        This method can be called after parsing to add multi-brand detection capabilities
        without breaking existing single-brand workflows.
        
        Args:
            parse_result: The ParseResult to analyze for multi-brand content
            
        Returns:
            Updated ParseResult with brand detection information
        """
        try:
            from .multi_brand_detector import MultiBrandDetector
            from .brand_normalizer import BrandNormalizer
            
            # Initialize components
            brand_normalizer = BrandNormalizer()
            multi_brand_detector = MultiBrandDetector(brand_normalizer)
            
            # Perform brand detection
            brand_detection_result = multi_brand_detector.detect_brands(parse_result)
            
            # Update the ParseResult with brand detection information
            parse_result.brand_detection_result = brand_detection_result
            parse_result.is_multi_brand = brand_detection_result.is_multi_brand
            
            return parse_result
            
        except Exception as e:
            # If multi-brand detection fails, log the error but don't break the workflow
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Multi-brand detection failed: {e}")
            
            # Return original ParseResult unchanged
            return parse_result

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def _find_best_data_sheet(self, file_path: str, available_sheets: list) -> str:
        """
        Find the best sheet that contains creator data.
        
        Criteria:
        1. Has USERNAME and BRAND columns
        2. Has most rows with valid data
        3. Not a config/summary sheet
        
        Returns:
            Sheet name with highest score, or first sheet if none found
        """
        import pandas as pd
        import sys
        
        skip_keywords = ['sow', 'link sku', 'gmv team', 'hitung', 'shopee', 'database', 'summary', 'config']
        
        # Filter out non-data sheets
        candidate_sheets = [
            s for s in available_sheets
            if not any(kw in s.lower() for kw in skip_keywords)
        ]
        
        if not candidate_sheets:
            candidate_sheets = available_sheets
        
        print(f"[PARSER] Evaluating {len(candidate_sheets)} candidate sheets", file=sys.stderr)
        
        # Score each sheet (limit to first 20 rows for speed)
        best_sheet = None
        best_score = 0
        
        for sheet_name in candidate_sheets:
            try:
                # Only read first 20 rows for quick evaluation
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
                
                # Quick check for USERNAME and BRAND in first 10 rows
                has_username = False
                has_brand = False
                header_row = -1
                
                for i in range(min(10, len(df))):
                    row = df.iloc[i]
                    row_str = ' '.join([str(v).upper() for v in row.values if pd.notna(v)])
                    if 'USERNAME' in row_str:
                        has_username = True
                        header_row = i
                    if 'BRAND' in row_str:
                        has_brand = True
                    if has_username and has_brand:
                        break
                
                # Calculate score
                score = 0
                if has_username:
                    score += 10
                if has_brand:
                    score += 10
                if has_username and has_brand:
                    # Estimate data rows (actual count will be done during full parse)
                    score += 50
                
                if score > best_score:
                    best_score = score
                    best_sheet = sheet_name
                    
            except Exception as e:
                print(f"[PARSER] Skip sheet '{sheet_name}': {e}", file=sys.stderr)
                continue
        
        result = best_sheet or (available_sheets[0] if available_sheets else 'Deal')
        print(f"[PARSER] Selected: '{result}' (score={best_score})", file=sys.stderr)
        return result

    def parse(self, file_path: str) -> ParseResult:
        """
        Method utama. Membaca file spreadsheet dan mengembalikan ParseResult.

        - .xlsx / .xls / .ods : baca sheet "Deal" dan "Belum Deal" secara terpisah
        - .csv                 : semua baris masuk deal_rows, non_deal_rows = []
        Raises UnsupportedFormatError jika ekstensi tidak didukung.
        """
        ext = Path(file_path).suffix.lower()
        if ext not in self.SUPPORTED_FORMATS:
            raise UnsupportedFormatError(
                f"Format '{ext}' tidak didukung. "
                f"Format yang diterima: {', '.join(self.SUPPORTED_FORMATS)}"
            )

        errors: list[str] = []
        deal_rows: list[CreatorRow] = []
        non_deal_rows: list[CreatorRow] = []
        detected_columns: list[str] = []
        template_info: Optional[TemplateInfo] = None

        if ext == '.csv':
            try:
                df = self._read_file(file_path)
                header_idx = self.detect_header(df)
                df = self._promote_header(df, header_idx)
                detected_columns = list(df.columns)
                standard_mapping, custom_columns = self._build_mapping(df.columns)
                deal_rows = self._df_to_rows(df, standard_mapping, custom_columns)
            except UnsupportedFormatError:
                raise
            except Exception as exc:
                errors.append(f"Gagal membaca file CSV: {exc}")
        else:
            # Excel / ODS — coba baca sheet Deal, fallback ke sheet pertama
            try:
                # Cek apakah sheet "Deal" ada
                import openpyxl
                wb_check = openpyxl.load_workbook(file_path, read_only=True)
                available_sheets = wb_check.sheetnames
                wb_check.close()

                deal_sheet = self.SHEET_DEAL if self.SHEET_DEAL in available_sheets else None
                non_deal_sheet = self.SHEET_NON_DEAL if self.SHEET_NON_DEAL in available_sheets else None

                # Jika sheet Deal tidak ada, cari sheet terbaik dengan data creator
                if not deal_sheet:
                    deal_sheet = self._find_best_data_sheet(file_path, available_sheets)
                    import sys
                    print(f"[PARSER] Auto-selected sheet: '{deal_sheet}' from {len(available_sheets)} available sheets", file=sys.stderr)

                df_deal = self._read_sheet(file_path, deal_sheet)
                header_idx = self.detect_header(df_deal)
                df_deal = self._promote_header(df_deal, header_idx)
                detected_columns = list(df_deal.columns)
                
                # Deteksi template setelah header di-promote
                template_info = self.template_detector.detect_template(df_deal, file_path)
                import sys
                print(f"[PARSER] Detected template: {template_info.template_type} "
                      f"(confidence: {template_info.confidence:.2f}, "
                      f"multi-video: {template_info.multi_video_support})", file=sys.stderr)
                
                standard_mapping, custom_columns = self._build_mapping(df_deal.columns)

                # Ambil data_start_row yang akurat (termasuk sub-header rows)
                actual_data_start = df_deal.attrs.get('data_start_row', header_idx + 1)

                # Ekstrak cell notes SEBELUM filtering rows
                video_links_by_row = {}
                if ext in ('.xlsx', '.xls'):
                    # Skip video extraction during initial parse for speed
                    # Video links will be extracted during report generation
                    import os
                    skip_video_extraction = os.environ.get('SKIP_VIDEO_EXTRACTION', 'false').lower() == 'true'
                    
                    if not skip_video_extraction:
                        # Gunakan template info untuk menentukan strategi ekstraksi
                        parsing_recommendations = self.template_detector.get_parsing_recommendations(template_info)
                        video_columns = parsing_recommendations['video_columns_to_scan']
                        
                        if not video_columns:
                            # Fallback ke deteksi otomatis jika template tidak punya rekomendasi
                            video_columns = []
                            for col in detected_columns:
                                col_upper = col.upper()
                                if any(kw in col_upper for kw in ['LINK VIDEO', 'VIDEO LINK', 'UPDATE VT', 'FOLLOWUP', 'VT']):
                                    video_columns.append(col)
                        
                        print(f"[PARSER] Extracting video links from {len(video_columns)} columns", file=sys.stderr)

                        # Ekstrak dari semua kolom video dan gabungkan hasilnya
                        max_videos = parsing_recommendations.get('max_videos_per_creator', 5)
                        for video_col in video_columns:
                            # Pass actual_data_start - 1 sebagai header_row_idx
                            # karena _extract_cell_notes menghitung _dsr = header_row_idx + 2
                            col_links = self._extract_cell_notes(
                                file_path, deal_sheet, video_col,
                                actual_data_start - 1  # -1 karena _dsr = idx + 2
                            )
                            # Merge hasil: tambahkan link baru ke setiap row
                            for row_idx, links in col_links.items():
                                if row_idx not in video_links_by_row:
                                    video_links_by_row[row_idx] = []
                                # Hindari duplikat dan batasi jumlah video per creator
                                for link in links:
                                    if link not in video_links_by_row[row_idx] and len(video_links_by_row[row_idx]) < max_videos:
                                        video_links_by_row[row_idx].append(link)
                        
                        print(f"[PARSER] Extracted video links for {len(video_links_by_row)} creators", file=sys.stderr)
                    else:
                        print(f"[PARSER] Skipping video extraction for faster upload", file=sys.stderr)
                
                # Convert DataFrame ke CreatorRow dengan video links
                deal_rows = self._df_to_rows_with_links(
                    df_deal, 
                    standard_mapping, 
                    custom_columns,
                    video_links_by_row
                )
                
                # Log hasil akhir per creator
                import sys
                for row in deal_rows:
                    if row.video_links:
                        print(f"[PARSER] Creator {row.username}: {len(row.video_links)} video links {row.video_links}", file=sys.stderr)
            except Exception as exc:
                errors.append(f"Gagal membaca sheet Deal: {exc}")

            # Baca sheet Belum Deal (opsional)
            if non_deal_sheet:
                try:
                    df_non = self._read_sheet(file_path, non_deal_sheet)
                    header_idx_non = self.detect_header(df_non)
                    df_non = self._promote_header(df_non, header_idx_non)
                    if not detected_columns:
                        detected_columns = list(df_non.columns)
                    std_map_non, cust_col_non = self._build_mapping(df_non.columns)
                    non_deal_rows = self._df_to_rows(df_non, std_map_non, cust_col_non)
                except Exception:
                    non_deal_rows = []

        return ParseResult(
            deal_rows=deal_rows,
            non_deal_rows=non_deal_rows,
            detected_columns=detected_columns,
            errors=errors,
            template_info=template_info,
            brand_detection_result=None,  # Will be populated by multi-brand detection
            is_multi_brand=False,  # Will be set by multi-brand detection
        )

    def detect_header(self, df: pd.DataFrame) -> int:
        """
        Deteksi baris header otomatis.
        Cari baris yang mengandung paling banyak kolom standar (case-insensitive).
        Return index baris (0-based). Default 0 jika tidak ada yang cocok.
        """
        import re
        # Normalize standard columns dengan strip karakter non-alphanumeric
        standard_upper = {
            ' '.join(re.sub(r'[^\w\s/]', '', col).upper().replace('\n', ' ').split())
            for col in self.STANDARD_COLUMNS
        }
        best_idx = 0
        best_count = 0

        for i, row in df.iterrows():
            row_values = {
                ' '.join(re.sub(r'[^\w\s/]', '', str(v)).strip().upper().replace('\n', ' ').split())
                for v in row.values if pd.notna(v)
            }
            count = len(row_values & standard_upper)
            if count > best_count:
                best_count = count
                best_idx = int(i)

        # Jika tidak ada baris yang cocok dengan kolom standar (best_count == 0),
        # cek apakah baris pertama sudah berisi data (bukan header)
        # dengan melihat apakah kolom ke-2 (USERNAME) berisi string non-header
        if best_count == 0:
            return -1  # Tidak ada header — data mulai dari row 0

        return best_idx

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _read_sheet(self, file_path: str, sheet_name: str) -> pd.DataFrame:
        """Baca satu sheet dari file Excel/ODS, return DataFrame mentah (tanpa header)."""
        ext = Path(file_path).suffix.lower()
        engine = self._engine_for(ext)
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine=engine)
        return df
    def _read_file(self, file_path: str) -> pd.DataFrame:
        """Baca file CSV, return DataFrame mentah (tanpa header)."""
        df = pd.read_csv(file_path, header=None, dtype=str)
        return df

    def _promote_header(self, df: pd.DataFrame, header_idx: int) -> pd.DataFrame:
        """
        Jadikan baris ke-header_idx sebagai nama kolom.
        Handle multi-row header: merge baris header yang berdekatan.
        Jika header_idx == -1, gunakan nama kolom default berdasarkan posisi.
        """
        if header_idx == -1:
            standard_cols = [
                'TGL DEAL', 'USERNAME', 'LINK ACC', 'FOLLS', 'CONTACT', 'BRAND', 'PIC',
                'AVG GMV/MONTH', 'UPDATE', 'RESPON', 'SPEED', 'RESULT', 'STATUS SEMPEL',
                'LINK VIDEO', 'TOTAL VT', 'NOTE DEAL'
            ]
            cols = list(df.columns)
            new_cols = []
            for i, col in enumerate(cols):
                if i < len(standard_cols):
                    new_cols.append(standard_cols[i])
                else:
                    new_cols.append(str(col))
            df = df.copy()
            df.columns = new_cols
            df = df.dropna(how='all')
            return df

        # Ambil baris header utama
        main_header = [str(v).strip() if pd.notna(v) else '' for v in df.iloc[header_idx]]

        # Cek apakah ada sub-header di baris berikutnya (multi-row header)
        # Sub-header: baris setelah header yang masih berisi nama kolom (bukan data)
        import re
        sub_header_rows = []
        for sub_idx in range(header_idx + 1, min(header_idx + 3, len(df))):
            sub_row = df.iloc[sub_idx]
            # Cek apakah baris ini adalah sub-header (mengandung nama kolom standar)
            sub_values = [str(v).strip() if pd.notna(v) else '' for v in sub_row]
            sub_norm = {
                ' '.join(re.sub(r'[^\w\s/]', '', v).upper().replace('\n', ' ').split())
                for v in sub_values if v
            }

            # Hitung berapa banyak nilai yang cocok dengan kolom standar
            std_upper = {
                ' '.join(re.sub(r'[^\w\s/]', '', c).upper().replace('\n', ' ').split())
                for c in self.STANDARD_COLUMNS
            }
            matches = len(sub_norm & std_upper)

            # Juga cek apakah ada keyword kolom
            has_col_keyword = any(
                kw in v.upper()
                for v in sub_values if v
                for kw in ['VT', 'GMV', 'NOTE', 'USERNAME', 'JUMLAH', 'TOTAL', 'UPDATE', 'SAMPEL', 'LINK']
            )

            # Cek apakah baris ini adalah data (bukan header)
            # Data biasanya berisi tanggal, URL, angka — bukan nama kolom
            has_data_pattern = any(
                any(pat in str(v) for pat in ['http', 'tiktok.com', 'February', 'March', 'January', 'April'])
                for v in sub_values if v
            )

            if has_data_pattern:
                break  # Ini baris data, bukan sub-header

            if matches > 0 or has_col_keyword:
                sub_header_rows.append(sub_values)
            else:
                break

        # Merge header: untuk kolom yang kosong di main_header, isi dari sub-header
        # Juga ganti kolom non-standar dengan sub-header yang lebih informatif
        merged_header = list(main_header)

        # Build set kolom standar untuk lookup cepat
        import re
        std_norm_set = {
            ' '.join(re.sub(r'[^\w\s/]', '', c).upper().replace('\n', ' ').split())
            for c in self.STANDARD_COLUMNS
        }
        std_norm_set.update({
            ' '.join(re.sub(r'[^\w\s/]', '', k).upper().replace('\n', ' ').split())
            for k in self._COLUMN_TO_FIELD
        })

        def _is_standard(col_name: str) -> bool:
            # Strip karakter non-alphanumeric kecuali spasi dan /
            import re
            cleaned = re.sub(r'[^\w\s/]', '', col_name)
            n = ' '.join(cleaned.upper().replace('\n', ' ').split())
            return n in std_norm_set

        for sub_values in sub_header_rows:
            for i, val in enumerate(sub_values):
                if not val:
                    continue
                if i < len(merged_header):
                    current = merged_header[i]
                    if not current:
                        # Kolom kosong → isi dari sub
                        merged_header[i] = val
                    elif not _is_standard(current) and _is_standard(val):
                        # Current bukan standar, sub adalah standar → ganti
                        merged_header[i] = val
                    elif not _is_standard(current) and not _is_standard(val):
                        # Keduanya bukan standar → pakai sub jika lebih informatif
                        pass  # keep current
                else:
                    # Kolom baru dari sub-header
                    if val:
                        merged_header.append(val)

        # Deduplicate kolom
        seen_cols: dict[str, int] = {}
        final_header = []
        for col in merged_header:
            if col in seen_cols:
                seen_cols[col] += 1
                final_header.append(f"{col}_{seen_cols[col]}")
            else:
                seen_cols[col] = 0
                final_header.append(col)

        # Tentukan baris data mulai dari mana
        data_start = header_idx + 1 + len(sub_header_rows)

        df_data = df.iloc[data_start:].copy()

        # Sesuaikan jumlah kolom
        n_cols = len(df_data.columns)
        n_header = len(final_header)
        if n_header < n_cols:
            final_header.extend([f'_col_{i}' for i in range(n_header, n_cols)])
        elif n_header > n_cols:
            final_header = final_header[:n_cols]

        df_data.columns = final_header
        df_data = df_data.reset_index(drop=True)
        df_data = df_data.dropna(how='all')
        # Simpan data_start sebagai attribute untuk dipakai _extract_cell_notes
        df_data.attrs['data_start_row'] = data_start
        return df_data

    def _build_mapping(self, columns) -> tuple[dict[str, str], list[str]]:
        """
        Bangun dua struktur dari daftar kolom DataFrame:
        - standard_mapping: {nama_kolom_df: field_CreatorRow}
        - custom_columns: kolom yang tidak ada di STANDARD_COLUMNS
        """
        # Normalize helper
        def _norm(s: str) -> str:
            import re
            # Strip karakter non-alphanumeric kecuali spasi dan /
            cleaned = re.sub(r'[^\w\s/]', '', str(s))
            return ' '.join(cleaned.strip().replace('\n', ' ').upper().split())

        # Build lookup dari STANDARD_COLUMNS
        standard_upper_map = {}
        for col in self.STANDARD_COLUMNS:
            standard_upper_map[_norm(col)] = col

        # Juga build dari _COLUMN_TO_FIELD keys langsung
        for col in self._COLUMN_TO_FIELD:
            standard_upper_map[_norm(col)] = col

        standard_mapping: dict[str, str] = {}
        custom_columns: list[str] = []
        mapped_fields: set[str] = set()

        for col in columns:
            col_norm = _norm(col)
            if not col_norm:
                continue

            # 1. Exact match
            matched_std = standard_upper_map.get(col_norm)

            # 2. Match tanpa spasi di sekitar /
            if not matched_std:
                col_no_slash = col_norm.replace('/ ', '/').replace(' /', '/')
                matched_std = standard_upper_map.get(col_no_slash)

            # 3. Fuzzy: cek apakah kolom mengandung keyword standar
            if not matched_std:
                matched_std = self._fuzzy_match_column(col_norm, standard_upper_map)

            if matched_std:
                field_name = self._COLUMN_TO_FIELD.get(matched_std)
                if field_name and field_name not in mapped_fields:
                    standard_mapping[col] = field_name
                    mapped_fields.add(field_name)
                elif not field_name:
                    # Kolom dikenali tapi tidak ada field mapping → custom
                    if str(col).strip():
                        custom_columns.append(col)
            else:
                if str(col).strip():
                    custom_columns.append(col)

        return standard_mapping, custom_columns

    def _fuzzy_match_column(self, col_norm: str, standard_map: dict) -> str | None:
        """
        Fuzzy matching untuk kolom yang namanya sedikit berbeda.
        Strategi:
        1. Kolom mengandung keyword standar (substring match)
        2. Keyword standar mengandung kolom (reverse substring)
        3. Similarity score sederhana
        """
        # Keyword penting dan field-nya
        KEYWORD_MAP = {
            'USERNAME': 'USERNAME',
            'TANGGAL': 'TANGGAL',
            'LINK ACC': 'LINK ACC',
            'LINK AKUN': 'LINK ACC',
            'FOLLS': 'FOLLS',
            'FOLLOWER': 'FOLLS',
            'CONTACT': 'CONTACT',
            'KONTAK': 'CONTACT',
            'PIC': 'PIC',
            'APPROVED': 'APPROVED BY RARA',
            'AVG GMV': 'AVG GMV/MONTH',
            'GMV/MONTH': 'AVG GMV/MONTH',
            'GMV MONTH': 'AVG GMV/MONTH',
            'GMV PER BULAN': 'GMV PERBULAN AFTER JOIN',
            'GMV PERBULAN': 'GMV PERBULAN AFTER JOIN',
            'UPDATE': 'UPDATE',
            'RESPON': 'RESPON SPEED',
            'RESULT': 'RESULT',
            'SAMPEL': 'SAMPEL GRATIS',
            'SAMPLE': 'SAMPEL GRATIS',
            'NOTE DEAL': 'NOTE DEAL',
            'NOTE RARA': 'NOTE DARI RARA',
            'TOTAL VT': 'TOTAL VT',
            'JUMLAH VT': 'TOTAL VT',
            'LINK VIDEO': 'LINK VIDEO',
            'VIDEO LINK': 'VIDEO LINK',
            'FOLLOWUP': 'UPDATE VT FOLLOWUP',
        }

        # Special case: kolom yang mengandung "FOLLOWUP" dan ("UPDATE" atau "VIDEO" atau "VT")
        # kemungkinan besar adalah kolom link video
        if 'FOLLOWUP' in col_norm:
            if any(kw in col_norm for kw in ['UPDATE', 'VIDEO', 'VT', 'LINK']):
                return 'UPDATE VT FOLLOWUP'

        for keyword, std_col in KEYWORD_MAP.items():
            if keyword in col_norm:
                if std_col in standard_map:
                    return std_col
                # Cari di _COLUMN_TO_FIELD
                if std_col in self._COLUMN_TO_FIELD:
                    return std_col

        # Reverse: apakah keyword standar ada di dalam nama kolom
        for std_norm, std_col in standard_map.items():
            if len(std_norm) >= 4 and std_norm in col_norm:
                return std_col
            if len(col_norm) >= 4 and col_norm in std_norm:
                return std_col

        return None

    def _df_to_rows(
        self,
        df: pd.DataFrame,
        standard_mapping: dict[str, str],
        custom_columns: list[str],
    ) -> list[CreatorRow]:
        """Convert DataFrame ke list[CreatorRow], skip baris yang bukan data creator."""
        return self._df_to_rows_with_links(df, standard_mapping, custom_columns, {})

    def _df_to_rows_with_links(
        self,
        df: pd.DataFrame,
        standard_mapping: dict[str, str],
        custom_columns: list[str],
        video_links_by_row: dict[int, list[str]],
    ) -> list[CreatorRow]:
        """Convert DataFrame ke list[CreatorRow] dengan video links dari cell notes."""
        rows: list[CreatorRow] = []

        # Kata-kata yang menandakan baris adalah label section, bukan data
        _SECTION_LABELS = {
            'MARET', 'FEBRUARI', 'JANUARI', 'APRIL', 'MEI', 'JUNI', 'JULI',
            'AGUSTUS', 'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER',
            'BUAT BESOK', 'FOLLOWUP', '[FOLLOWUP]', 'NEXT', 'LANJUTAN',
        }

        for df_idx, pandas_row in df.iterrows():
            kwargs: dict = {}

            # Isi field standar
            for col, field_name in standard_mapping.items():
                raw = pandas_row.get(col)
                value = self._clean_value(raw)

                if value is not None:
                    if field_name in self._INT_FIELDS:
                        value = self._to_int(value)
                    elif field_name in self._FLOAT_FIELDS:
                        value = self._to_float(value)
                    else:
                        value = str(value) if value is not None else None

                kwargs[field_name] = value

            # Isi custom_fields
            custom: dict[str, str] = {}
            for col in custom_columns:
                raw = pandas_row.get(col)
                val = self._clean_value(raw)
                custom[col] = str(val) if val is not None else ''

            kwargs['custom_fields'] = custom
            row = CreatorRow(**kwargs)

            # Skip baris yang bukan data creator:
            # 1. Tidak ada username DAN tidak ada link_acc
            # 2. Username adalah label section (MARET, FEBRUARI, dll)
            # 3. Username adalah URL kosong (https://www.tiktok.com/@?shop_region=ID)
            username = (row.username or '').strip()
            link = (row.link_acc or '').strip()

            if not username and not link:
                continue

            # Skip label section
            if username.upper() in _SECTION_LABELS:
                continue

            # Skip URL kosong (username kosong di URL)
            if link in ('https://www.tiktok.com/@?shop_region=ID', 'https://www.tiktok.com/@'):
                continue
            if username == '@' or username == '':
                continue

            # Tambahkan video links dari cell notes jika ada
            # df_idx adalah index di DataFrame (setelah promote_header)
            # video_links_by_row menggunakan index relatif dari Excel
            if df_idx in video_links_by_row:
                row.video_links = video_links_by_row[df_idx]

            rows.append(row)

        return rows

    # ------------------------------------------------------------------
    # Utility
    # ------------------------------------------------------------------

    def _extract_cell_notes(
        self,
        file_path: str,
        sheet_name: str,
        col_name: str,
        header_row_idx: int,
    ) -> dict[int, list[str]]:
        """Ekstrak hyperlinks dan data validation TikTok dari sheet Excel."""
        import re as _re
        import zipfile as _zipfile
        import xml.etree.ElementTree as _ET
        import openpyxl as _openpyxl
        from openpyxl.utils import get_column_letter as _gcl

        try:
            _wb = _openpyxl.load_workbook(file_path, read_only=True)
            if sheet_name not in _wb.sheetnames:
                _wb.close()
                return {}
            _sheet_idx = _wb.sheetnames.index(sheet_name) + 1
            _ws = list(_wb.worksheets)[_sheet_idx - 1]
            _target_col = 'M'
            for _row in _ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 1):
                for _cell in _row:
                    if _cell.value:
                        _n = ' '.join(_re.sub(r'[^\w\s/]', '', str(_cell.value)).strip().upper().replace('\n', ' ').split())
                        _cn = ' '.join(_re.sub(r'[^\w\s/]', '', col_name).upper().replace('\n', ' ').split())
                        if _n == _cn:
                            _target_col = _gcl(_cell.column)
                            break
            _wb.close()

            # Pattern untuk mendeteksi TikTok URLs - lebih comprehensive
            _url_re = _re.compile(r'https?://[^\s,;"<>\n\r]+tiktok[^\s,;"<>\n\r]*', _re.IGNORECASE)
            _result: dict[int, list[str]] = {}
            _dsr = max(header_row_idx + 2, 1)  # data start row (1-indexed)

            def _clean_url(url: str) -> str:
                """Bersihkan URL dari suffix tanggal/komentar yang ikut terbawa."""
                url = url.strip()
                # Hapus whitespace dan karakter kontrol
                url = _re.sub(r'[\s\r\n\t]+', '', url)
                # Potong di koma pertama jika ada (tanggal sering dipisah koma)
                if ',' in url:
                    url = url.split(',')[0].strip()
                
                # Potong jika ada URL lain yang langsung bersambung
                # Cari pattern: ...?param=valuehttps:// atau ...video/123456https://
                next_url = _re.search(r'(https?://)', url[8:])  # Skip first https://
                if next_url:
                    url = url[:8 + next_url.start()]
                
                # Hapus trailing punctuation
                url = url.rstrip('.,;:!?_')
                
                # Pastikan URL valid
                if not url.startswith('http'):
                    return ''
                if 'tiktok' not in url.lower():
                    return ''
                
                # Hapus trailing underscore atau karakter aneh di akhir
                # yang mungkin dari pemisahan yang salah
                url = _re.sub(r'[_\-]+$', '', url)
                
                return url

            def _extract_urls_from_text(text: str) -> list[str]:
                """Ekstrak semua TikTok URLs dari text, handling multiple separators."""
                if not text:
                    return []
                
                # Split by common separators: newline, comma, semicolon, space
                # Tapi tetap preserve URL yang utuh
                urls = []
                
                # Method 1: Regex extraction (paling reliable)
                # Improved pattern to handle URLs that are concatenated
                found_urls = _url_re.findall(text)
                for url in found_urls:
                    cleaned = _clean_url(url)
                    if cleaned and cleaned not in urls:
                        urls.append(cleaned)
                
                # Method 2: Handle concatenated URLs (no separator between them)
                # Look for pattern: ...video/123456https://... or ...?param=valuehttps://...
                if 'httpshttps' in text.lower() or text.count('https://') > len(urls):
                    # Split by 'https://' and reconstruct URLs
                    parts = text.split('https://')
                    for i, part in enumerate(parts):
                        if i == 0 and not part.startswith('http'):
                            continue  # Skip first part if it doesn't start with http
                        
                        # Reconstruct URL
                        url = 'https://' + part
                        
                        # Find where this URL ends (next https:// or http:// or end of string)
                        # Look for the next URL start
                        next_url_match = _re.search(r'(https?://)', url[8:])  # Skip the first https://
                        if next_url_match:
                            # Cut at the next URL
                            url = url[:8 + next_url_match.start()]
                        
                        cleaned = _clean_url(url)
                        if cleaned and 'tiktok' in cleaned.lower() and cleaned not in urls:
                            urls.append(cleaned)
                
                # Method 3: Split by newline/comma dan coba parse each part
                if not urls:
                    parts = _re.split(r'[\n\r,;]+', text)
                    for part in parts:
                        part = part.strip()
                        if 'tiktok' in part.lower() and part.startswith('http'):
                            cleaned = _clean_url(part)
                            if cleaned and cleaned not in urls:
                                urls.append(cleaned)
                
                return urls

            with _zipfile.ZipFile(file_path) as _z:
                _sc = _z.read(f'xl/worksheets/sheet{_sheet_idx}.xml').decode('utf-8')

                # Method 1: Hyperlinks dari _rels
                _rels_path = f'xl/worksheets/_rels/sheet{_sheet_idx}.xml.rels'
                if _rels_path in _z.namelist():
                    _rels_root = _ET.fromstring(_z.read(_rels_path).decode('utf-8'))
                    _rid_map: dict[str, str] = {}
                    for _rel in _rels_root:
                        _rid = _rel.get('Id', '')
                        _tgt = _rel.get('Target', '')
                        if _rid and _tgt and 'tiktok' in _tgt.lower():
                            _rid_map[_rid] = _tgt
                    if _rid_map:
                        for _m in _re.finditer(r'<hyperlink\b([^/]*)/>', _sc, _re.IGNORECASE):
                            _attrs = _m.group(1)
                            _rm = _re.search(r'ref="([A-Z]+)(\d+)"', _attrs, _re.IGNORECASE)
                            _im = _re.search(r'r:id="(rId\d+)"', _attrs, _re.IGNORECASE)
                            if _rm and _im:
                                _cl = _rm.group(1).upper()
                                _rn = int(_rm.group(2))
                                _rid = _im.group(1)
                                if _cl == _target_col.upper() and _rid in _rid_map:
                                    _ri = _rn - _dsr
                                    if _ri >= 0:
                                        _cleaned = _clean_url(_rid_map[_rid])
                                        if _cleaned:
                                            if _ri not in _result:
                                                _result[_ri] = []
                                            if _cleaned not in _result[_ri]:
                                                _result[_ri].append(_cleaned)

                # Method 2: Data Validation dropdowns
                # Format: <dataValidation sqref="AE192"><formula1>"username,https://vt.tiktok.com/xxx/"</formula1>
                _dv_pat = _re.compile(
                    r'<dataValidation[^>]+sqref="([^"]+)"[^>]*>.*?<formula1>(.*?)</formula1>',
                    _re.DOTALL
                )
                for _m in _dv_pat.finditer(_sc):
                    _sqref = _m.group(1)
                    _formula = _m.group(2).replace('&quot;', '"').replace('&amp;', '&')
                    
                    # Extract all URLs from formula using improved method
                    _urls = _extract_urls_from_text(_formula)
                    
                    if _urls:
                        for _rn_m in _re.finditer(r'[A-Z]+(\d+)', _sqref):
                            _rn = int(_rn_m.group(1))
                            _ri = _rn - _dsr
                            if _ri >= 0:
                                if _ri not in _result:
                                    _result[_ri] = []
                                for url in _urls:
                                    if url not in _result[_ri]:
                                        _result[_ri].append(url)

                # Method 3: Cell values (untuk cell yang berisi plain text URLs)
                # Parse cell values dari sharedStrings.xml
                try:
                    _ss_path = 'xl/sharedStrings.xml'
                    if _ss_path in _z.namelist():
                        _ss_content = _z.read(_ss_path).decode('utf-8')
                        _ss_root = _ET.fromstring(_ss_content)
                        
                        # Extract all text from shared strings
                        _shared_strings = []
                        for _si in _ss_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                            _text_parts = []
                            for _t in _si.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t'):
                                if _t.text:
                                    _text_parts.append(_t.text)
                            _shared_strings.append(''.join(_text_parts))
                        
                        # Now parse cell values and match with shared strings
                        _cell_pat = _re.compile(
                            rf'<c r="({_target_col}\d+)"[^>]*t="s"[^>]*><v>(\d+)</v></c>',
                            _re.IGNORECASE
                        )
                        for _m in _cell_pat.finditer(_sc):
                            _cell_ref = _m.group(1)
                            _ss_idx = int(_m.group(2))
                            
                            if _ss_idx < len(_shared_strings):
                                _cell_text = _shared_strings[_ss_idx]
                                _urls = _extract_urls_from_text(_cell_text)
                                
                                if _urls:
                                    _rn = int(_re.search(r'\d+', _cell_ref).group())
                                    _ri = _rn - _dsr
                                    if _ri >= 0:
                                        if _ri not in _result:
                                            _result[_ri] = []
                                        for url in _urls:
                                            if url not in _result[_ri]:
                                                _result[_ri].append(url)
                except Exception as _e:
                    # Jika gagal parse sharedStrings, skip method ini
                    import sys
                    print(f"[PARSER] Warning: Failed to parse sharedStrings: {_e}", file=sys.stderr)

            return _result
        except Exception as _e:
            import sys
            print(f"[PARSER] Error in _extract_cell_notes: {_e}", file=sys.stderr)
            return {}

    @staticmethod
    def _engine_for(ext: str) -> str:
        if ext in ('.xlsx',):
            return 'openpyxl'
        if ext in ('.xls',):
            return 'xlrd'
        if ext in ('.ods',):
            return 'odf'
        return 'openpyxl'

    @staticmethod
    def _clean_value(raw):
        """Return None jika NaN/None, else return raw."""
        if raw is None:
            return None
        try:
            if isinstance(raw, float) and math.isnan(raw):
                return None
        except TypeError:
            pass
        if isinstance(raw, str) and raw.strip() == '':
            return None
        return raw

    @staticmethod
    def _to_int(value) -> Optional[int]:
        try:
            return int(float(str(value).replace(',', '').replace('.', '')))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _to_float(value) -> Optional[float]:
        try:
            s = str(value).strip()
            # Handle format Indonesia: Rp7,1JT, Rp1,5M, Rp500RB, dll
            s = s.replace('Rp', '').replace('RP', '').replace(' ', '')
            multiplier = 1
            if s.upper().endswith('JT'):
                multiplier = 1_000_000
                s = s[:-2]
            elif s.upper().endswith('M'):
                multiplier = 1_000_000
                s = s[:-1]
            elif s.upper().endswith('RB'):
                multiplier = 1_000
                s = s[:-2]
            elif s.upper().endswith('K'):
                multiplier = 1_000
                s = s[:-1]
            # Ganti koma desimal ke titik
            s = s.replace(',', '.')
            # Jika ada lebih dari satu titik, hapus titik ribuan
            parts = s.split('.')
            if len(parts) > 2:
                s = ''.join(parts[:-1]) + '.' + parts[-1]
            return float(s) * multiplier
        except (ValueError, TypeError):
            return None
